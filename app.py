import sys
sys.path.append('/Users/moriyasuyuto/Library/Python/3.8/lib/python/site-packages')

# requestsのインポート
import requests 
# Beautiful Soupのインポート
from bs4 import BeautifulSoup

import time
import datetime

from selenium import webdriver
from selenium.webdriver.chrome.options import Options

import openpyxl
import re

import json

import urllib.parse

excel_path = "../../Desktop/pritimes_0608_0621.xlsx"

json_data = json.load(open('./data.json', 'r'))

start_date = '2022-06-08' + 'T00:00:00+09:00'

end_date = '2022-06-21' + 'T23:59:59+09:00'

def getQuery(keyword):
    query = json_data[keyword]['query']
    return query

def getUrl(keyword):
    url = 'https://prtimes.jp/main/action.php?run=html&page=searchkey&search_word='
    url_quote = url + urllib.parse.quote(getQuery(keyword))
    return url_quote

# def getCount(keyword):
#     count = json_data[keyword]['count']
#     return count

def illegal_char_remover(data):
    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]|[\x00-\x1f\x7f-\x9f]|[\uffff]')
    """Remove ILLEGAL CHARACTER."""
    if isinstance(data, str):
        return ILLEGAL_CHARACTERS_RE.sub("", data)
    else:
        return data    

def collectData(keyword):
    print("operation starts. keyword: " + keyword)
    options = Options()
    # ヘッドレスモードを有効にする
    options.add_argument('--headless')
    # ChromeのWebDriverオブジェクトを作成する
    driver = webdriver.Chrome(options=options, executable_path="/Applications/chromedriver")
    driver.get(getUrl(keyword))
    for i in range(100):
        html = driver.page_source.encode('utf-8')
        soup = BeautifulSoup(html, 'html.parser')
        time_tags = soup.find_all(class_='list-article__time')
        last_time_tag = time_tags[len(time_tags) - 1]

        # ISO 8601 を datetime オブジェクトに変換
        start_date_dt = datetime.datetime.fromisoformat(start_date)
        last_time_tag_dt = datetime.datetime.fromisoformat(last_time_tag.get('datetime').replace('0900', '09:00'))
        # 最後の日付が start_time より小さくなったら処理をやめる
        if (start_date_dt > last_time_tag_dt): break

        driver.find_element_by_class_name('js-list-article-more-button').click()
        time.sleep(2)

    #テスト URL, 会社名, 記事タイトル
    a_tags= soup.find_all(class_='list-article__link')
    h3_tags=soup.find_all(class_='list-article__title')
    company_link_tags=soup.find_all(class_='list-article__company-name-link')

    blog_links=[]
    for a_tag in a_tags:
        #記事URLをblog_linksにappend
        blog_links.append('https://prtimes.jp' + a_tag.get('href'))
    print('article links acquired')

    blog_titles=[]
    for h3_tag in h3_tags:
        #記事タイトルをblog_titlesにappend
        nobr = h3_tag.text.replace('\n', '')
        nospace = nobr.replace(' ', '')
        blog_titles.append(nospace)
    print('article titles acquired')

    blog_companies=[]
    for company_link_tag in company_link_tags:
        #会社名をblog_companiesにappend
        nobr_company = company_link_tag.text.replace('\n', '')
        nospace_company = nobr_company.replace(' ', '')
        blog_companies.append(nospace_company)
    print('article companies acquired')

    updated_times=[]
    tel_numbers=[]
    for blog_link in blog_links:
        r = requests.get(blog_link)
        soup = BeautifulSoup(r.text, 'html.parser')
        updated_times.append(soup.time.text)
        tel_numbers.append(soup.find_all(class_='body-information')[3].text.replace('\n', '').replace(' ', ''))
    print('updated times acquired')
    print('tel numbers acquired')

    # Excelにデータを書き込み
    wb = openpyxl.load_workbook(excel_path)
    wb.create_sheet(title=getQuery(keyword))
    ws = wb[getQuery(keyword)]

    ws.cell(row=1, column=2, value=illegal_char_remover('会社名'))
    ws.cell(row=1, column=4, value=illegal_char_remover('記事タイトル'))
    ws.cell(row=1, column=5, value=illegal_char_remover('記事URL'))
    ws.cell(row=1, column=6, value=illegal_char_remover('電話番号'))
    ws.cell(row=1, column=7, value=illegal_char_remover('更新日時'))
    ws.cell(row=1, column=8, value=illegal_char_remover('取得KW'))
    for i in range(len(blog_links)):
        ws.cell(row=i+2, column=2, value=illegal_char_remover(blog_companies[i]))
        ws.cell(row=i+2, column=4, value=illegal_char_remover(blog_titles[i]))
        ws.cell(row=i+2, column=5, value=illegal_char_remover(blog_links[i]))
        ws.cell(row=i+2, column=6, value=illegal_char_remover(tel_numbers[i]))
        ws.cell(row=i+2, column=7, value=illegal_char_remover(updated_times[i]))
        ws.cell(row=i+2, column=8, value=illegal_char_remover(getQuery(keyword)))
    wb.save(excel_path)
    print('data written in excel')
    print('operation ended successfully. keyword: ' + keyword + '\n')

for data in json_data:
    collectData(data)
    time.sleep(1)
